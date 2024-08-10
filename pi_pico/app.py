import openpyxl
import os
import serial
from datetime import datetime

desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
excel_file = os.path.join(desktop, 'sensor.xlsx')

if not os.path.exists(excel_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sensor Data"
    ws.append(["Date", "Time", "Sensor 1", "Sensor 2", "Sensor 3"])
    wb.save(excel_file)
else:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

try:
    ser = serial.Serial('/dev/ttyACM0', 115200, timeout=1)
except serial.SerialException as e:
    print(f"SerialException: {e}")
    exit()

def save_data(sensor1_value, sensor2_value, sensor3_value):
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")
    empty_row = ws.max_row + 1
    ws.cell(row=empty_row, column=1, value=date)
    ws.cell(row=empty_row, column=2, value=time_str)
    ws.cell(row=empty_row, column=3, value=sensor1_value)
    ws.cell(row=empty_row, column=4, value=sensor2_value)
    ws.cell(row=empty_row, column=5, value=sensor3_value)
    wb.save(excel_file)

while True:
    try:
        data = ser.readline().decode().strip()
        if data:
            sensor_values = data.split(',')
            if len(sensor_values) == 3:
                sensor1_value, sensor2_value, sensor3_value = sensor_values
                save_data(sensor1_value, sensor2_value, sensor3_value)
                print(f"Data saved at {datetime.now().strftime('%H:%M:%S')}")
    except serial.SerialException as e:
        print(f"SerialException: {e}")
        break
    except Exception as e:
        print(f"Unexpected error: {e}")
        break
